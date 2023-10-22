#
# MIT License
#
# Copyright (c) 2023 c0fec0de
#

"""
Add Home Assistant Data to Excel.
"""

import argparse
import re
import sys
from copy import copy
from datetime import datetime, timedelta
from pathlib import Path

import openpyxl
import requests

INVALID_NAME = "INVALID-NAME"
INVALID_TYPE = "INVALID-TYPE"
_RE_SPEC = re.compile(r"\A(?P<name>[^\|]*)(\|(?P<type>.*))?\Z")
_TYPEFUNCS = {
    "float": float,
    "int": int,
}


def _get_entity(host, port, token, name, timeout=60):
    url = f"http://{host}:{port}/api/states/{name}"
    headers = headers = {
        "Authorization": f"Bearer {token}",
        "content-type": "application/json",
    }
    response = requests.get(url, headers=headers, timeout=timeout)
    data = response.json() or {}
    return data.get("state", INVALID_NAME)


def _get_cell(data, host, port, token, spec):
    mat = _RE_SPEC.match(spec)
    name = mat["name"]
    type_ = mat["type"]
    try:
        value = data[name]
    except KeyError:
        value = _get_entity(host, port, token, name)

    if not type_:
        return value

    try:
        typefunc = _TYPEFUNCS[type_]
        try:
            return typefunc(value)
        except ValueError as exc:
            return str(exc)
    except KeyError:
        return INVALID_TYPE


def _add_row(sheet, host, port, token, specs, timestamp):
    # pylint: disable=too-many-arguments,too-many-locals
    data = {
        "time": timestamp.time(),
        "datetime": timestamp,
        "date": timestamp.date(),
    }

    # Search 'latest' marker
    for rowidx, row in enumerate(sheet.iter_rows(), 1):
        cell = row[0]
        comment = cell.comment
        if comment and comment.text.strip() == "latest":
            # add new row
            rowidx += 1
            sheet.insert_rows(rowidx)
            # Move 'latest' marker
            cell.comment = None
            sheet.cell(rowidx, 1).comment = comment
            break
    else:
        sheet.append([""])
        rowidx = sheet.max_row

    prev_row, row = sheet.iter_rows(min_row=rowidx - 1, max_row=rowidx)

    for spec, cell, prev_cell in zip(specs, row, prev_row):
        # style
        cell.font = copy(prev_cell.font)
        cell.border = copy(prev_cell.border)
        cell.fill = copy(prev_cell.fill)
        cell.number_format = copy(prev_cell.number_format)
        cell.protection = copy(prev_cell.protection)
        cell.alignment = copy(prev_cell.alignment)
        # value
        if spec:
            # retrieve value
            cell.value = _get_cell(data, host, port, token, spec)
        else:
            # replicate formulas
            if str(prev_cell.value).startswith("="):
                translator = openpyxl.formula.translate.Translator(prev_cell.value, origin=prev_cell.coordinate)
                sheet[cell.coordinate] = translator.translate_formula(cell.coordinate)

    # repair formulas behind
    for row in sheet.iter_rows(min_row=rowidx + 1):
        for cell in row:
            if str(cell.value).startswith("="):
                translator = openpyxl.formula.translate.Translator(cell.value, origin=cell.coordinate)
                sheet[cell.coordinate] = translator.translate_formula(row_delta=1)


def run(xlsxpath: Path, host: str, port: str, token: str, timestamp: datetime):
    """
    Run.

    Args:
        xlsxpath: Path to XLSX file

    Keyword Args:
        host: Home Assistant Host
        port: Home Assistant Port
        token: Home Assistant Token
        timestamp: Data Timestamp.
    """
    # Open Workbook, Update Sheets
    book = openpyxl.load_workbook(xlsxpath)
    book.iso_dates = True
    for sheet in book:
        try:
            specs = [cell.comment and cell.comment.text.strip() for cell in next(sheet.rows)]
        except StopIteration:
            continue
        if any(specs):
            _add_row(sheet, host, port, token, specs, timestamp)

    # Save Workbook
    book.save(xlsxpath)


def main(args=None):
    """Command Line Interface."""
    parser = argparse.ArgumentParser(
        prog="homeassistant2xls",
        description="Copy Data From Home-Assistant to XLSX files",
    )
    parser.add_argument("xlsx")  # positional argument
    parser.add_argument("--token", help="Home Assistant API Token. Mandatory.")
    parser.add_argument("--host", default="localhost", help="Home Assistant Port. 'localhost' by default.")
    parser.add_argument("--port", default="8123", help="Home Assistant Port. '8123' by default.")
    parser.add_argument("--timeoffset", help="Timestamp offset in minutes. Positive and negative numbers are allowed.")
    # importlib is not available in py37
    pyversion = (sys.version_info.major, sys.version_info.minor)
    if pyversion >= (3, 8):  # pragma: no cover
        # pylint: disable=import-outside-toplevel
        import importlib.metadata

        version = importlib.metadata.version("homeassistant2xlsx")
        parser.add_argument("--version", action="version", version=f"homeassistant2xlsx {version}")
    # parse
    args = parser.parse_args(args=args)

    timestamp = datetime.now()
    if args.timeoffset:
        timestamp += timedelta(minutes=int(args.timeoffset))

    run(Path(args.xlsx), args.host, args.port, args.token, timestamp)
